# Внешние зависимости
from typing import Dict, List
import os
from uuid import UUID
import subprocess
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from fastapi import HTTPException, status
# Внутренние модули
from web_app.src.schemas import FormRequest
from web_app.src.core import cfg


# Получаем валидные ячейки, которые нужно заполнить данными
def get_valid_cells() -> None:
    result: Dict[int, List[str]] = {}

    wb = load_workbook(f"{cfg.XLSX_DIR}/{cfg.TEMPLATE_XLSX}")

    try:
        for sheet_index in range(len(wb.worksheets)):
            ws = wb.worksheets[sheet_index]
            merged_ranges = ws.merged_cells.ranges

            if not merged_ranges:
                continue

            if result.get(sheet_index) is None:
                result[sheet_index] = []

            sorted_ranges = sorted(merged_ranges, key=lambda x: (x.min_row, x.min_col))

            for merged_range in sorted_ranges:
                min_row = merged_range.min_row
                min_col = merged_range.min_col
                max_row = merged_range.max_row

                if max_row < cfg.FILTER_SETTINGS[sheet_index]["skip"]:
                    continue

                cell_value = ws.cell(row=min_row, column=min_col).value

                if cell_value is not None and str(cell_value).strip() != "":
                    continue

                min_col_letter = get_column_letter(min_col)
                excel_format = f"{min_col_letter}{min_row}"

                if cfg.FILTER_SETTINGS[sheet_index]["ignore"] is not None and \
                        excel_format in cfg.FILTER_SETTINGS[sheet_index]["ignore"]:
                    continue

                result[sheet_index].append(excel_format)

    except Exception as err:
        cfg.logger.error(f"Error in get valid cells: {err}")

    finally:
        if wb:
            wb.close()

    cfg.VALID_CELLS = result


# Генерируем файл xlsx с данными
def create_xlsx_from_data(data: FormRequest, output_name: UUID) -> None:
    data_dict = data.model_dump()
    all_words = []
    for key, value in data_dict.items():
        if key in cfg.TRANSFER_WORDS:
            continue

        if value in (True, False):
            all_words.append("✓" if value else " ")
        else:
            all_words.append(value)

    all_words.append(data.surname_field)
    all_words.append(data.name_field)
    all_words.append(data.patronymic_field[:-1])
    all_words.append(data.citizenship_field + " ")
    all_words.append(data.birth_day)
    all_words.append(data.birth_month)
    all_words.append(data.birth_year)
    all_words.append("✓" if data.gender_male else " ")
    all_words.append("✓" if data.gender_female else " ")
    all_words.append(data.document_type[:10])
    all_words.append(data.document_series)
    all_words.append(data.document_number[:10])
    all_words.append(data.issue_day)
    all_words.append(data.issue_month)
    all_words.append(data.issue_year)
    all_words.append(data.expiry_day)
    all_words.append(data.expiry_month)
    all_words.append(data.expiry_year)
    all_words.append(data.prev_region)
    all_words.append(data.prev_city)
    all_words.append(data.prev_area)
    all_words.append(data.prev_road)
    all_words.append(data.prev_building_1)
    all_words.append(data.prev_building_2)
    all_words.append(data.prev_building_3)
    all_words.append(data.prev_room_1)
    all_words.append(data.prev_room_2)
    all_words.append(data.permission_doc_line_6)
    all_words.append(data.actual_region)
    all_words.append(data.actual_city)
    all_words.append(data.actual_area)
    all_words.append(data.actual_cadastral)
    all_words.append(data.live_day)
    all_words.append(data.live_month)
    all_words.append(data.live_year)

    for word in cfg.TRANSFER_WORDS:
        all_words.append(data_dict[word])

    all_words.append(data.surname_field)
    all_words.append(data.name_field)
    all_words.append(data.patronymic_field[:-1])
    all_words.append(data.host_name_org_line_1 + data.host_name_org_line_2[:-3])
    all_words.append(data.host_org_inn[:-1])
    all_words.append(" " * 8)

    data_for_cells = []
    for word in all_words:
        if word[0] == "~":
            data_for_cells.append(word[1:-1])
        else:
            data_for_cells.extend(word)

    wb = load_workbook(f"{cfg.XLSX_DIR}/{cfg.TEMPLATE_XLSX}")
    try:
        len_data_for_cells = len(data_for_cells)
        i = 0
        for sheet_index in range(len(wb.worksheets)):
            ws = wb.worksheets[sheet_index]

            if cfg.VALID_CELLS.get(sheet_index) is not None:
                for cell in cfg.VALID_CELLS[sheet_index]:
                    if i < len_data_for_cells:
                        letters = data_for_cells[i]
                        i += 1

                    else:
                        letters = " "

                    ws[cell] = letters

        output_file = f"{cfg.XLSX_DIR}/{output_name}.xlsx"
        wb.save(output_file)

    except Exception as err:
        cfg.logger.error(f"Error in create xlsx: {err}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Unexpected error")

    finally:
        if wb:
            wb.close()


# Конвертируем xlsx в pdf
def exel_to_pdf_libreoffice(filename: UUID) -> None:
    cmd = [
        "libreoffice",
        "--headless",
        "--convert-to", "pdf",
        "--outdir", cfg.PDF_DIR,
        f"{cfg.XLSX_DIR}/{filename}.xlsx"
    ]

    try:
        subprocess.run(cmd, capture_output=True, text=True, check=True)
        cfg.logger.info("Конвертация успешно завершена")

    except subprocess.CalledProcessError as err:
        cfg.logger.error(f"Ошибка конвертации: {err}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Unexpected error")

    except Exception as err:
        cfg.logger.error(f"Неизвестная ошибка: {err}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Unexpected error")


# Безопасно удаляем файл с обработкой ошибок
def delete_file_safe(filename: UUID) -> None:
    try:
        file_path = f"{cfg.XLSX_DIR}/{filename}.xlsx"
        os.remove(file_path)
        cfg.logger.info(f"Файл '{file_path}' успешно удален")

    except FileNotFoundError:
        cfg.logger.error(f"Файл '{file_path}' не найден")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Unexpected error")

    except PermissionError:
        cfg.logger.error(f"Нет прав для удаления файла '{file_path}'")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Unexpected error")

    except OSError as e:
        cfg.logger.error(f"Ошибка при удалении файла '{file_path}': {e}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Unexpected error")


def run_generate_pdf(data: FormRequest, output_name: UUID):
    create_xlsx_from_data(data=data, output_name=output_name)
    exel_to_pdf_libreoffice(filename=output_name)
    delete_file_safe(filename=output_name)
